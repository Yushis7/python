import openpyxl
import pandas as pd
# import dataframe as df
from datetime import datetime

# Load the workbook
wb = openpyxl.load_workbook('라벨출력(sample).xlsx')

# Select the source and target sheets
sheet_a = wb['석간수기라밸']
sheet_b = wb['Sheet1']

# Define the range of cells to move
start_row, start_col = 1, 1  # Change as needed
end_row, end_col = 10, 10    # Change as needed

# Iterate through the specified range and move cells
for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell_value = sheet_a.cell(row=row, column=col).value
        sheet_b.cell(row=row, column=col).value = cell_value
        sheet_a.cell(row=row, column=col).value = None

# Save the workbook
wb.save('라벨출력(sample).xlsx')
