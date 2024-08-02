from openpyxl import load_workbook
from openpyxl import Workbook

# 엑셀 파일을 불러옵니다.
file_path = "C:/주간/파이썬/begin1.xlsx"
workbook = load_workbook(filename=file_path)
sheet = workbook.active  # 활성 시트를 사용합니다. 필요한 경우 sheet 이름을 명시할 수도 있습니다.

# 새로운 워크북과 시트를 만듭니다.
new_workbook = Workbook()
new_sheet = new_workbook.active

# 헤더를 복사합니다. (A열 외에도 다른 열이 있을 경우를 대비)
for col in sheet.iter_cols(min_row=1, max_row=1):
    for cell in col:
        new_sheet[cell.coordinate] = cell.value

# A열에서 '사슴'인 행만 복사합니다.
row_index = 2  # 새 시트에 쓸 행의 인덱스 (헤더 이후이므로 2부터 시작)
for row in sheet.iter_rows(min_row=2):  # 헤더는 제외하고 2번째 행부터 시작
    if row[0].value == '사슴':  # A열의 값을 체크
        for col_index, cell in enumerate(row, 1):
            new_sheet.cell(row=row_index, column=col_index, value=cell.value)
        row_index += 1

# 결과를 새로운 엑셀 파일로 저장합니다.
output_file_path = 'filtered_excel_file.xlsx'
new_workbook.save(output_file_path)

print(f"Filtered data has been saved to {output_file_path}")