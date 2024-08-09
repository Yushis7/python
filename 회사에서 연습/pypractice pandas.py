import os
import sys
import pandas as pd
import openpyxl
from datetime import datetime

setting = pd.read_excel('settings.xlsx', sheet_name='Sheet1')

origin_file = pd.read_excel('settings.xlsx', sheet_name='Sheet2')
origin_f = origin_file.values[0][0]
file_open = pd.read_excel(origin_f)

columns_list = setting[['추출 파일 열', '원래 열']].values.tolist()

df = pd.DataFrame()

for x, y in columns_list:
    df[x] = file_open[y]


#datatime.today()는 오늘 날짜이고  strftime('%Y%m%d')는 연도, 월, 일을 가져오는 함수이다.
df.to_excel('excel_완료_' + datetime.today().strftime('%Y%m%d') + '.xls', index=False, engine='openpyxl')





workbook = openpyxl.load_workbook("test2.xlsx")
first_sheet = workbook.active
first_sheet.title = "원본데이터"

second_sheet = workbook.create_sheet(title="복사 데이터")

for row in first_sheet.iter_rows(min_row=2, values_only=True):
    name,email = row[0],row[3]
    second_sheet.append([name,email])

workbook.save("test02_copy.xlsx")

