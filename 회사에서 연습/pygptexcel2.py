from openpyxl import load_workbook

# 엑셀 파일을 불러옵니다.
file_path = 'your_excel_file.xlsx'
workbook = load_workbook(filename=file_path)
sheet = workbook.active  # 활성 시트를 사용합니다. 필요한 경우 sheet 이름을 명시할 수도 있습니다.

# 삭제할 행을 기록합니다.
rows_to_delete = []

# 첫 번째 행(헤더)은 건너뛰고 두 번째 행부터 확인합니다.
for row in sheet.iter_rows(min_row=2):
    if row[0].value != '사슴':  # A열의 값이 '사슴'이 아닌 경우
        rows_to_delete.append(row[0].row)

# 뒤에서부터 행을 삭제합니다 (행 번호가 변경되는 것을 방지하기 위해)
for row_index in reversed(rows_to_delete):
    sheet.delete_rows(row_index, 1)

# 변경된 엑셀 파일을 저장합니다.
output_file_path = 'filtered_excel_file.xlsx'
workbook.save(output_file_path)

print(f"Filtered data has been saved to {output_file_path}")