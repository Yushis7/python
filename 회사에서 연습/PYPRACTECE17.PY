from openpyxl import Workbook ,load_workbook
 
# 새로운 워크북 생성
workbook = Workbook()



 
# 활성 시트 선택
sheet = workbook.active
 
# 데이터 입력
sheet.append(['ID', 'Fruit', 'Value'])
sheet.append([1, 'apple', 3.14])
sheet.append([2, 'banana', 2.71])
sheet.append([3, 'cherry', 1.41])
 
# 'Fruit' 열 값이 'banana'인 행 삭제
for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, values_only=True):
    if row[0:3] == 'banana':
        sheet.delete_rows(row[0])
 
# 엑셀 파일 저장
workbook.save('지정 필터 삭제.xlsx')

